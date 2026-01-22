"""Excel export functionality using openpyxl."""

from pathlib import Path
from typing import Optional, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database.models import RFQ, Part, Tool, Material, Machine


# Styles
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

WARNING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def set_column_widths(ws, widths: dict):
    """Set column widths for a worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def export_rfq_to_excel(
    rfq: RFQ,
    parts: List[Part],
    tools: List[Tool],
    output_path: str | Path,
    include_calculations: bool = True,
    include_sanity_checks: bool = True
) -> str:
    """Export RFQ data to Excel file.

    Args:
        rfq: RFQ model instance
        parts: List of Part model instances
        tools: List of Tool model instances
        output_path: Path to save Excel file
        include_calculations: Include calculation details
        include_sanity_checks: Include sanity check results

    Returns:
        Path to created Excel file
    """
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, rfq, parts, tools)

    # Parts sheet
    ws_parts = wb.create_sheet("Parts")
    _write_parts_sheet(ws_parts, parts, include_calculations)

    # Tools sheet
    ws_tools = wb.create_sheet("Tools")
    _write_tools_sheet(ws_tools, tools, include_calculations, include_sanity_checks)

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return str(output_path)


def _write_summary_sheet(ws, rfq: RFQ, parts: List[Part], tools: List[Tool]):
    """Write summary information."""
    ws['A1'] = "RFQ Summary"
    ws['A1'].font = Font(bold=True, size=14)

    ws['A3'] = "RFQ Name:"
    ws['B3'] = rfq.name
    ws['A4'] = "Customer:"
    ws['B4'] = rfq.customer or "-"
    ws['A5'] = "Status:"
    ws['B5'] = rfq.status
    ws['A6'] = "Created:"
    ws['B6'] = rfq.created_date.strftime('%Y-%m-%d') if rfq.created_date else "-"

    ws['A8'] = "Parts:"
    ws['B8'] = len(parts)
    ws['A9'] = "Tools:"
    ws['B9'] = len(tools)

    # Calculate totals
    total_demand_sop = sum(p.demand_sop or 0 for p in parts)
    total_demand_peak = sum(p.demand_peak or 0 for p in parts)
    total_tool_price = sum(t.price_enquiry or t.price_estimated or 0 for t in tools)

    ws['A11'] = "Total Demand (SOP):"
    ws['B11'] = total_demand_sop
    ws['A12'] = "Total Demand (Peak):"
    ws['B12'] = total_demand_peak
    ws['A13'] = "Total Tool Investment:"
    ws['B13'] = f"€ {total_tool_price:,.2f}" if total_tool_price else "-"

    ws['A15'] = "Notes:"
    ws['A16'] = rfq.notes or "-"
    ws.merge_cells('A16:D20')

    set_column_widths(ws, {1: 20, 2: 30, 3: 15, 4: 15})


def _write_parts_sheet(ws, parts: List[Part], include_calculations: bool):
    """Write parts data."""
    headers = [
        "Part Name", "Part Number", "Material",
        "Weight (g)", "Volume (cm³)", "Proj. Area (cm²)", "Wall (mm)",
        "Demand SOP", "Demand EAOP", "Demand Peak", "Lifetime Vol.",
        "Cycle Time (s)"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    for row, part in enumerate(parts, 2):
        ws.cell(row=row, column=1, value=part.name)
        ws.cell(row=row, column=2, value=part.part_number or "-")
        ws.cell(row=row, column=3, value=part.material.short_name if part.material else "-")
        ws.cell(row=row, column=4, value=part.weight_g)
        ws.cell(row=row, column=5, value=part.volume_cm3)
        ws.cell(row=row, column=6, value=part.projected_area_cm2)
        ws.cell(row=row, column=7, value=part.wall_thickness_mm)
        ws.cell(row=row, column=8, value=part.demand_sop)
        ws.cell(row=row, column=9, value=part.demand_eaop)
        ws.cell(row=row, column=10, value=part.demand_peak)
        ws.cell(row=row, column=11, value=part.parts_over_runtime)
        ws.cell(row=row, column=12, value=part.cycle_time_s)

        # Apply borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    set_column_widths(ws, {
        1: 25, 2: 15, 3: 12,
        4: 10, 5: 12, 6: 14, 7: 10,
        8: 12, 9: 12, 10: 12, 11: 12,
        12: 12
    })


def _write_tools_sheet(ws, tools: List[Tool], include_calculations: bool, include_sanity_checks: bool):
    """Write tools data."""
    headers = [
        "Tool Name", "Type", "Cavities",
        "Injection System", "Surface", "Sliders", "Lifters",
        "Dimensions (LxWxH)", "Clamping (kN)", "Machine",
        "Fits?", "Complexity",
        "Supplier", "Country",
        "Price Enquiry", "Price Estimate", "Notes"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    for row, tool in enumerate(tools, 2):
        ws.cell(row=row, column=1, value=tool.name)
        ws.cell(row=row, column=2, value=tool.tool_type)
        ws.cell(row=row, column=3, value=tool.cavities)
        ws.cell(row=row, column=4, value=tool.injection_system)
        ws.cell(row=row, column=5, value=tool.surface_finish)
        ws.cell(row=row, column=6, value=tool.sliders_count)
        ws.cell(row=row, column=7, value=tool.lifters_count)

        # Dimensions
        dims = []
        if tool.tool_length_mm:
            dims.append(f"{tool.tool_length_mm:.0f}")
        if tool.tool_width_mm:
            dims.append(f"{tool.tool_width_mm:.0f}")
        if tool.tool_height_mm:
            dims.append(f"{tool.tool_height_mm:.0f}")
        ws.cell(row=row, column=8, value=" x ".join(dims) if dims else "-")

        ws.cell(row=row, column=9, value=tool.estimated_clamping_force_kn)
        ws.cell(row=row, column=10, value=tool.machine.name if tool.machine else "-")

        # Fit status with coloring
        fit_cell = ws.cell(row=row, column=11)
        if tool.fits_machine is None:
            fit_cell.value = "?"
        elif tool.fits_machine:
            fit_cell.value = "Yes"
        else:
            fit_cell.value = "NO"
            fit_cell.fill = ERROR_FILL

        ws.cell(row=row, column=12, value=tool.complexity_rating or "-")
        ws.cell(row=row, column=13, value=tool.supplier_name or "-")
        ws.cell(row=row, column=14, value=tool.supplier_country or "-")

        # Prices
        if tool.price_enquiry:
            ws.cell(row=row, column=15, value=f"€ {tool.price_enquiry:,.2f}")
        else:
            ws.cell(row=row, column=15, value="-")

        if tool.price_estimated:
            ws.cell(row=row, column=16, value=f"€ {tool.price_estimated:,.2f}")
        else:
            ws.cell(row=row, column=16, value="-")

        ws.cell(row=row, column=17, value=tool.notes or "")

        # Apply borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    set_column_widths(ws, {
        1: 25, 2: 10, 3: 10,
        4: 15, 5: 12, 6: 8, 7: 8,
        8: 18, 9: 12, 10: 15,
        11: 8, 12: 10,
        13: 20, 14: 12,
        15: 15, 16: 15, 17: 30
    })


def export_existing_tools_to_excel(
    tools: List,  # List of ExistingTool
    output_path: str | Path
) -> str:
    """Export existing tools reference data to Excel.

    Args:
        tools: List of ExistingTool model instances
        output_path: Path to save Excel file

    Returns:
        Path to created Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Existing Tools"

    headers = [
        "Name", "Description", "Part Type",
        "Complexity", "Cavities", "Sliders", "Lifters",
        "Surface", "Injection System",
        "Dimensions (LxWxH)", "Steel Weight (kg)",
        "Supplier", "Country", "Price", "Date",
        "Issues", "Lessons Learned", "Tags"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    for row, tool in enumerate(tools, 2):
        ws.cell(row=row, column=1, value=tool.name)
        ws.cell(row=row, column=2, value=tool.description or "-")
        ws.cell(row=row, column=3, value=tool.part_type or "-")
        ws.cell(row=row, column=4, value=tool.complexity_rating or "-")
        ws.cell(row=row, column=5, value=tool.cavities)
        ws.cell(row=row, column=6, value=tool.sliders_count)
        ws.cell(row=row, column=7, value=tool.lifters_count)
        ws.cell(row=row, column=8, value=tool.surface_finish or "-")
        ws.cell(row=row, column=9, value=tool.injection_system or "-")

        # Dimensions
        dims = []
        if tool.tool_length_mm:
            dims.append(f"{tool.tool_length_mm:.0f}")
        if tool.tool_width_mm:
            dims.append(f"{tool.tool_width_mm:.0f}")
        if tool.tool_height_mm:
            dims.append(f"{tool.tool_height_mm:.0f}")
        ws.cell(row=row, column=10, value=" x ".join(dims) if dims else "-")

        ws.cell(row=row, column=11, value=tool.steel_weight_kg or "-")
        ws.cell(row=row, column=12, value=tool.supplier_name or "-")
        ws.cell(row=row, column=13, value=tool.supplier_country or "-")

        if tool.actual_price:
            ws.cell(row=row, column=14, value=f"{tool.currency} {tool.actual_price:,.2f}")
        else:
            ws.cell(row=row, column=14, value="-")

        ws.cell(row=row, column=15, value=tool.price_date.strftime('%Y-%m-%d') if tool.price_date else "-")
        ws.cell(row=row, column=16, value=tool.issues or "-")
        ws.cell(row=row, column=17, value=tool.lessons_learned or "-")
        ws.cell(row=row, column=18, value=tool.tags or "-")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return str(output_path)
